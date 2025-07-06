from flask import Flask, render_template_string, request, redirect, url_for, session, send_file
import json
import os
from datetime import datetime
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = 'supersecretkey'  # Oturum için gerekli

ADMIN_PASSWORD = 'admin123'
DATA_FILE = 'puantaj_data.json'
CALISAN_FILE = 'calisanlar.json'
SANTIYE_FILE = 'santiyeler.json'
PUANTAJ_FILE = 'puantajlar.json'
TASERON_FILE = 'taseronlar.json'
PERF_FILE = 'performans.json'

BOOTSTRAP = '''
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
body, .form-control, .form-select, .btn, table, th, td {
  font-size: 1.15rem !important;
}
h1, h2, h3, h4, h5, h6, .navbar-brand, .card-title {
  font-weight: bold !important;
  letter-spacing: 0.5px;
}
table th, table td {
  font-weight: 500;
}
</style>
'''

HEADER = '''
<nav class="navbar navbar-expand-lg navbar-dark bg-primary mb-4">
  <div class="container-fluid">
    <span class="navbar-brand mb-0 h1">DENSARA PERSONEL TAKİP</span>
    {% if session.logged_in %}
    <a class="btn btn-light" href="/logout">Çıkış Yap</a>
    {% endif %}
  </div>
</nav>
'''

# --- Yardımcı Fonksiyonlar ---
def load_data():
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_calisanlar():
    if not os.path.exists(CALISAN_FILE):
        return []
    with open(CALISAN_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_calisanlar(data):
    with open(CALISAN_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_santiyeler():
    if not os.path.exists(SANTIYE_FILE):
        return []
    with open(SANTIYE_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_santiyeler(data):
    with open(SANTIYE_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_puantajlar():
    if not os.path.exists(PUANTAJ_FILE):
        return []
    with open(PUANTAJ_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_puantajlar(data):
    with open(PUANTAJ_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_taseronlar():
    if not os.path.exists(TASERON_FILE):
        return []
    with open(TASERON_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_taseronlar(data):
    with open(TASERON_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_performans():
    if not os.path.exists(PERF_FILE):
        return []
    with open(PERF_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_performans(data):
    with open(PERF_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# --- HTML Şablonları ---
login_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:400px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="card-title text-center mb-4">Yönetici Girişi</h3>
      <form method="post">
        <div class="mb-3">
          <input type="password" class="form-control" name="password" placeholder="Şifre" required>
        </div>
        <button type="submit" class="btn btn-primary w-100">Giriş Yap</button>
        {% if error %}<p class="text-danger mt-3">{{ error }}</p>{% endif %}
      </form>
    </div>
  </div>
</div>
'''

menu_page = BOOTSTRAP + HEADER + '''
<div class="container">
  <div class="row justify-content-center">
    <div class="col-md-6">
      <div class="card shadow">
        <div class="card-body text-center">
          <h2 class="mb-4">Ana Menü</h2>
          <a href="/bordro" class="btn btn-outline-danger w-100 mb-2">Bordro/Maaş Raporu</a>
          <a href="/toplu_bordro" class="btn btn-outline-danger w-100 mb-2">Toplu Bordro</a>
          <a href="/santiye_rapor" class="btn btn-outline-dark w-100 mb-2">Şantiye Raporları</a>
          <a href="/taseron_rapor" class="btn btn-outline-dark w-100 mb-2">Taşeron Raporu</a>
          <a href="/performanslar" class="btn btn-outline-success w-100 mb-2">Performans Takibi</a>
          <a href="/performans_rapor" class="btn btn-outline-info w-100 mb-2">Performans Raporu</a>
          <a href="/sgk_rapor" class="btn btn-outline-warning w-100 mb-2">SGK Bildirgesi</a>
          <a href="/devamsizlik_rapor" class="btn btn-outline-primary w-100 mb-2">Devamsızlık/İzin/Fazla Mesai Raporu</a>
          <a href="/puantajlar" class="btn btn-outline-secondary w-100 mb-2">Günlük Puantaj</a>
          <a href="/kayitlar" class="btn btn-outline-primary w-100 mb-2">Giriş/Çıkış Kayıtları</a>
          <a href="/ozel" class="btn btn-outline-success w-100 mb-2">İzin/Vardiya/Fazla Mesai</a>
          <a href="/rapor" class="btn btn-outline-warning w-100 mb-2">Raporlama (Excel İndir)</a>
          <a href="/calisanlar" class="btn btn-outline-info w-100 mb-2">Çalışan Yönetimi</a>
          <a href="/santiyeler" class="btn btn-outline-dark w-100 mb-2">Şantiye Yönetimi</a>
          <a href="/taseronlar" class="btn btn-outline-secondary w-100 mb-2">Taşeron Yönetimi</a>
        </div>
      </div>
    </div>
  </div>
</div>
'''

kayitlar_page = BOOTSTRAP + HEADER + '''
<div class="container">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h3>Giriş/Çıkış Kayıtları</h3>
    <div>
      <a href="/kayit_ekle" class="btn btn-primary">Yeni Kayıt Ekle</a>
      <a href="/menu" class="btn btn-secondary">Ana Menü</a>
    </div>
  </div>
  <div class="table-responsive">
    <table class="table table-striped table-bordered">
      <thead class="table-primary">
        <tr><th>Tarih</th><th>Saat</th><th>Tür</th><th>Açıklama</th></tr>
      </thead>
      <tbody>
      {% for k in kayitlar %}
        <tr><td>{{k['tarih']}}</td><td>{{k['saat']}}</td><td>{{k['tur']}}</td><td>{{k['aciklama']}}</td></tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
</div>
'''

kayit_ekle_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">Yeni Giriş/Çıkış Kaydı</h3>
      <form method="post">
        <div class="mb-3">
          <label>Tarih:</label>
          <input type="date" class="form-control" name="tarih" required value="{{today}}">
        </div>
        <div class="mb-3">
          <label>Saat:</label>
          <input type="time" class="form-control" name="saat" required>
        </div>
        <div class="mb-3">
          <label>Tür:</label>
          <select class="form-select" name="tur">
            <option value="Giriş">Giriş</option>
            <option value="Çıkış">Çıkış</option>
          </select>
        </div>
        <div class="mb-3">
          <label>Açıklama:</label>
          <input type="text" class="form-control" name="aciklama">
        </div>
        <button type="submit" class="btn btn-success w-100">Kaydet</button>
      </form>
    </div>
  </div>
</div>
'''

ozel_page = BOOTSTRAP + HEADER + '''
<div class="container">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h3>İzin/Vardiya/Fazla Mesai Kayıtları</h3>
    <div>
      <a href="/ozel_ekle" class="btn btn-success">Yeni Kayıt Ekle</a>
      <a href="/menu" class="btn btn-secondary">Ana Menü</a>
    </div>
  </div>
  <div class="table-responsive">
    <table class="table table-striped table-bordered">
      <thead class="table-success">
        <tr><th>Tarih</th><th>Saat</th><th>Tür</th><th>Açıklama</th></tr>
      </thead>
      <tbody>
      {% for k in kayitlar %}
        <tr><td>{{k['tarih']}}</td><td>{{k['saat']}}</td><td>{{k['tur']}}</td><td>{{k['aciklama']}}</td></tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
</div>
'''

ozel_ekle_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">Yeni İzin/Vardiya/Fazla Mesai Kaydı</h3>
      <form method="post">
        <div class="mb-3">
          <label>Tarih:</label>
          <input type="date" class="form-control" name="tarih" required value="{{today}}">
        </div>
        <div class="mb-3">
          <label>Saat:</label>
          <input type="time" class="form-control" name="saat" required>
        </div>
        <div class="mb-3">
          <label>Tür:</label>
          <select class="form-select" name="tur">
            <option value="İzin">İzin</option>
            <option value="Vardiya">Vardiya</option>
            <option value="Fazla Mesai">Fazla Mesai</option>
          </select>
        </div>
        <div class="mb-3">
          <label>Açıklama:</label>
          <input type="text" class="form-control" name="aciklama">
        </div>
        <button type="submit" class="btn btn-success w-100">Kaydet</button>
      </form>
    </div>
  </div>
</div>
'''

calisan_list_page = BOOTSTRAP + HEADER + '''
<div class="container">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h3>Çalışan Listesi</h3>
    <a href="/calisan_ekle" class="btn btn-primary">Yeni Çalışan Ekle</a>
  </div>
  <div class="table-responsive">
    <table class="table table-striped table-bordered">
      <thead class="table-info">
        <tr>
          <th>Ad</th><th>Soyad</th><th>TC</th><th>SGK No</th><th>Departman</th><th>Görev</th><th>Ustalık</th><th>Vardiya</th><th>Şantiye</th><th>Giriş</th><th>Ayrılış</th><th>İşlem</th>
        </tr>
      </thead>
      <tbody>
      {% for c in calisanlar %}
        <tr>
          <td>{{c['ad']}}</td><td>{{c['soyad']}}</td><td>{{c['tc']}}</td><td>{{c['sgk']}}</td><td>{{c['departman']}}</td><td>{{c['gorev']}}</td><td>{{c['ustalik']}}</td><td>{{c['vardiya']}}</td><td>{{c['santiye']}}</td><td>{{c['giris']}}</td><td>{{c['ayrilis']}}</td>
          <td>
            <a href="/calisan_duzenle/{{loop.index0}}" class="btn btn-sm btn-warning">Düzenle</a>
            <a href="/calisan_sil/{{loop.index0}}" class="btn btn-sm btn-danger" onclick="return confirm('Silinsin mi?')">Sil</a>
          </td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
  <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
</div>
'''

calisan_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:600px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">{{baslik}}</h3>
      {% if santiye_list|length == 0 %}
        <div class="alert alert-warning">Önce <a href='/santiyeler'>şantiye ekleyin</a>.</div>
      {% endif %}
      <form method="post">
        <div class="row">
          <div class="col-md-6 mb-3"><label>Ad:</label><input type="text" class="form-control" name="ad" required value="{{c.ad}}"></div>
          <div class="col-md-6 mb-3"><label>Soyad:</label><input type="text" class="form-control" name="soyad" required value="{{c.soyad}}"></div>
        </div>
        <div class="row">
          <div class="col-md-6 mb-3"><label>TC:</label><input type="text" class="form-control" name="tc" required value="{{c.tc}}"></div>
          <div class="col-md-6 mb-3"><label>SGK No:</label><input type="text" class="form-control" name="sgk" value="{{c.sgk}}"></div>
        </div>
        <div class="row">
          <div class="col-md-6 mb-3"><label>Departman:</label><input type="text" class="form-control" name="departman" value="{{c.departman}}"></div>
          <div class="col-md-6 mb-3"><label>Görev:</label><input type="text" class="form-control" name="gorev" value="{{c.gorev}}"></div>
        </div>
        <div class="row">
          <div class="col-md-6 mb-3"><label>Ustalık Seviyesi:</label><input type="text" class="form-control" name="ustalik" value="{{c.ustalik}}"></div>
          <div class="col-md-6 mb-3"><label>Vardiya:</label><input type="text" class="form-control" name="vardiya" value="{{c.vardiya}}"></div>
        </div>
        <div class="row">
          <div class="col-md-6 mb-3">
            <label>Taşeron:</label>
            <select class="form-select" name="taseron">
              <option value="">Taşeron seçiniz</option>
              {% for t in taseron_list %}
                <option value="{{t['ad']}}" {% if c.taseron==t['ad'] %}selected{% endif %}>{{t['ad']}}</option>
              {% endfor %}
            </select>
          </div>
          <div class="col-md-6 mb-3">
            <label>Şantiye:</label>
            <select class="form-select" name="santiye">
              <option value="">Şantiye seçiniz</option>
              {% for s in santiye_list %}
                <option value="{{s['ad']}}" {% if c.santiye==s['ad'] %}selected{% endif %}>{{s['ad']}}</option>
              {% endfor %}
            </select>
          </div>
        </div>
        <div class="row">
          <div class="col-md-6 mb-3"><label>İşe Giriş Tarihi:</label><input type="date" class="form-control" name="giris" value="{{c.giris}}"></div>
          <div class="col-md-6 mb-3"><label>Ayrılış Tarihi:</label><input type="date" class="form-control" name="ayrilis" value="{{c.ayrilis}}"></div>
        </div>
        <div class="row">
          <div class="col-md-4 mb-3"><label>Yevmiye (₺):</label><input type="number" step="0.01" class="form-control" name="yevmiye" value="{{c.yevmiye}}"></div>
          <div class="col-md-4 mb-3"><label>Fazla Mesai Ücreti (₺/saat):</label><input type="number" step="0.01" class="form-control" name="fm_ucret" value="{{c.fm_ucret}}"></div>
          <div class="col-md-4 mb-3"><label>Avans (₺):</label><input type="number" step="0.01" class="form-control" name="avans" value="{{c.avans}}"></div>
        </div>
        <div class="mb-3"><label>Kesinti (₺):</label><input type="number" step="0.01" class="form-control" name="kesinti" value="{{c.kesinti}}"></div>
        <button type="submit" class="btn btn-success w-100" {% if santiye_list|length == 0 %}disabled{% endif %}>Kaydet</button>
      </form>
      <a href="/calisanlar" class="btn btn-secondary mt-3">Geri</a>
    </div>
  </div>
</div>
'''

santiye_list_page = BOOTSTRAP + HEADER + '''
<div class="container">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h3>Şantiye Listesi</h3>
    <a href="/santiye_ekle" class="btn btn-primary">Yeni Şantiye Ekle</a>
  </div>
  <div class="table-responsive">
    <table class="table table-striped table-bordered">
      <thead class="table-info">
        <tr>
          <th>Şantiye Adı</th><th>Lokasyon</th><th>Proje Kodu</th><th>Yönetici</th><th>İşlem</th>
        </tr>
      </thead>
      <tbody>
      {% for s in santiyeler %}
        <tr>
          <td>{{s['ad']}}</td><td>{{s['lokasyon']}}</td><td>{{s['proje_kodu']}}</td><td>{{s['yonetici']}}</td>
          <td>
            <a href="/santiye_duzenle/{{loop.index0}}" class="btn btn-sm btn-warning">Düzenle</a>
            <a href="/santiye_sil/{{loop.index0}}" class="btn btn-sm btn-danger" onclick="return confirm('Silinsin mi?')">Sil</a>
          </td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
  <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
</div>
'''

santiye_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">{{baslik}}</h3>
      <form method="post">
        <div class="mb-3"><label>Şantiye Adı:</label><input type="text" class="form-control" name="ad" required value="{{s.ad}}"></div>
        <div class="mb-3"><label>Lokasyon:</label><input type="text" class="form-control" name="lokasyon" value="{{s.lokasyon}}"></div>
        <div class="mb-3"><label>Proje Kodu:</label><input type="text" class="form-control" name="proje_kodu" value="{{s.proje_kodu}}"></div>
        <div class="mb-3"><label>Yönetici:</label><input type="text" class="form-control" name="yonetici" value="{{s.yonetici}}"></div>
        <button type="submit" class="btn btn-success w-100">Kaydet</button>
      </form>
      <a href="/santiyeler" class="btn btn-secondary mt-3">Geri</a>
    </div>
  </div>
</div>
'''

puantaj_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:600px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">Günlük Puantaj Kaydı Ekle</h3>
      <form method="post">
        <div class="mb-3">
          <label>Çalışan:</label>
          <select class="form-select" name="calisan" required>
            <option value="">Çalışan seçiniz</option>
            {% for c in calisanlar %}
              <option value="{{c['ad']}} {{c['soyad']}}" {% if kayit.calisan==c['ad']+' '+c['soyad'] %}selected{% endif %}>{{c['ad']}} {{c['soyad']}}</option>
            {% endfor %}
          </select>
        </div>
        <div class="mb-3">
          <label>Şantiye:</label>
          <select class="form-select" name="santiye" required>
            <option value="">Şantiye seçiniz</option>
            {% for s in santiyeler %}
              <option value="{{s['ad']}}" {% if kayit.santiye==s['ad'] %}selected{% endif %}>{{s['ad']}}</option>
            {% endfor %}
          </select>
        </div>
        <div class="mb-3">
          <label>Tarih:</label>
          <input type="date" class="form-control" name="tarih" required value="{{kayit.tarih}}">
        </div>
        <div class="mb-3">
          <label>Tür:</label>
          <select class="form-select" name="tur" required>
            <option value="Giriş" {% if kayit.tur=='Giriş' %}selected{% endif %}>Giriş</option>
            <option value="Çıkış" {% if kayit.tur=='Çıkış' %}selected{% endif %}>Çıkış</option>
            <option value="İzin" {% if kayit.tur=='İzin' %}selected{% endif %}>İzin</option>
            <option value="Devamsızlık" {% if kayit.tur=='Devamsızlık' %}selected{% endif %}>Devamsızlık</option>
            <option value="Rapor" {% if kayit.tur=='Rapor' %}selected{% endif %}>Rapor</option>
            <option value="Resmi Tatil" {% if kayit.tur=='Resmi Tatil' %}selected{% endif %}>Resmi Tatil</option>
          </select>
        </div>
        <div class="mb-3">
          <label>Saat:</label>
          <input type="time" class="form-control" name="saat" value="{{kayit.saat}}">
        </div>
        <div class="mb-3">
          <label>Açıklama:</label>
          <input type="text" class="form-control" name="aciklama" value="{{kayit.aciklama}}">
        </div>
        <button type="submit" class="btn btn-success w-100">Kaydet</button>
      </form>
      <a href="/puantajlar" class="btn btn-secondary mt-3">Geri</a>
    </div>
  </div>
</div>
'''

puantaj_list_page = BOOTSTRAP + HEADER + '''
<div class="container">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h3>Günlük Puantaj Kayıtları</h3>
    <a href="/puantaj_ekle" class="btn btn-primary">Yeni Kayıt Ekle</a>
  </div>
  <div class="table-responsive">
    <table class="table table-striped table-bordered">
      <thead class="table-info">
        <tr>
          <th>Çalışan</th><th>Şantiye</th><th>Tarih</th><th>Tür</th><th>Saat</th><th>Açıklama</th>
        </tr>
      </thead>
      <tbody>
      {% for p in puantajlar %}
        <tr>
          <td>{{p['calisan']}}</td><td>{{p['santiye']}}</td><td>{{p['tarih']}}</td><td>{{p['tur']}}</td><td>{{p['saat']}}</td><td>{{p['aciklama']}}</td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
  <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
</div>
'''

taseron_list_page = BOOTSTRAP + HEADER + '''
<div class="container">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h3>Taşeron Listesi</h3>
    <a href="/taseron_ekle" class="btn btn-primary">Yeni Taşeron Ekle</a>
  </div>
  <div class="table-responsive">
    <table class="table table-striped table-bordered">
      <thead class="table-info">
        <tr>
          <th>Firma Adı</th><th>Yetkili</th><th>Telefon</th><th>İşlem</th>
        </tr>
      </thead>
      <tbody>
      {% for t in taseronlar %}
        <tr>
          <td>{{t['ad']}}</td><td>{{t['yetkili']}}</td><td>{{t['telefon']}}</td>
          <td>
            <a href="/taseron_duzenle/{{loop.index0}}" class="btn btn-sm btn-warning">Düzenle</a>
            <a href="/taseron_sil/{{loop.index0}}" class="btn btn-sm btn-danger" onclick="return confirm('Silinsin mi?')">Sil</a>
          </td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
  <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
</div>
'''

taseron_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">{{baslik}}</h3>
      <form method="post">
        <div class="mb-3"><label>Firma Adı:</label><input type="text" class="form-control" name="ad" required value="{{t.ad}}"></div>
        <div class="mb-3"><label>Yetkili:</label><input type="text" class="form-control" name="yetkili" value="{{t.yetkili}}"></div>
        <div class="mb-3"><label>Telefon:</label><input type="text" class="form-control" name="telefon" value="{{t.telefon}}"></div>
        <button type="submit" class="btn btn-success w-100">Kaydet</button>
      </form>
      <a href="/taseronlar" class="btn btn-secondary mt-3">Geri</a>
    </div>
  </div>
</div>
'''

perf_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:600px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">Performans Kaydı Ekle</h3>
      <form method="post">
        <div class="mb-3">
          <label>Çalışan:</label>
          <select class="form-select" name="calisan" required>
            <option value="">Çalışan seçiniz</option>
            {% for c in calisanlar %}
              <option value="{{c['ad']}} {{c['soyad']}}" {% if kayit.calisan==c['ad']+' '+c['soyad'] %}selected{% endif %}>{{c['ad']}} {{c['soyad']}}</option>
            {% endfor %}
          </select>
        </div>
        <div class="mb-3">
          <label>Şantiye:</label>
          <select class="form-select" name="santiye" required>
            <option value="">Şantiye seçiniz</option>
            {% for s in santiyeler %}
              <option value="{{s['ad']}}" {% if kayit.santiye==s['ad'] %}selected{% endif %}>{{s['ad']}}</option>
            {% endfor %}
          </select>
        </div>
        <div class="mb-3">
          <label>Tarih:</label>
          <input type="date" class="form-control" name="tarih" required value="{{kayit.tarih}}">
        </div>
        <div class="mb-3">
          <label>İş Türü:</label>
          <input type="text" class="form-control" name="is_turu" value="{{kayit.is_turu}}" required>
        </div>
        <div class="mb-3">
          <label>Miktar:</label>
          <input type="number" step="0.01" class="form-control" name="miktar" value="{{kayit.miktar}}" required>
        </div>
        <div class="mb-3">
          <label>Birim:</label>
          <input type="text" class="form-control" name="birim" value="{{kayit.birim}}" required>
        </div>
        <div class="mb-3">
          <label>Prim/Ödül (₺):</label>
          <input type="number" step="0.01" class="form-control" name="prim" value="{{kayit.prim}}">
        </div>
        <div class="mb-3">
          <label>Açıklama:</label>
          <input type="text" class="form-control" name="aciklama" value="{{kayit.aciklama}}">
        </div>
        <button type="submit" class="btn btn-success w-100">Kaydet</button>
      </form>
      <a href="/performanslar" class="btn btn-secondary mt-3">Geri</a>
    </div>
  </div>
</div>
'''

perf_list_page = BOOTSTRAP + HEADER + '''
<div class="container">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h3>Performans Kayıtları</h3>
    <a href="/performans_ekle" class="btn btn-primary">Yeni Kayıt Ekle</a>
  </div>
  <div class="table-responsive">
    <table class="table table-striped table-bordered">
      <thead class="table-info">
        <tr>
          <th>Çalışan</th><th>Şantiye</th><th>Tarih</th><th>İş Türü</th><th>Miktar</th><th>Birim</th><th>Prim/Ödül</th><th>Açıklama</th>
        </tr>
      </thead>
      <tbody>
      {% for p in performanslar %}
        <tr>
          <td>{{p['calisan']}}</td><td>{{p['santiye']}}</td><td>{{p['tarih']}}</td><td>{{p['is_turu']}}</td><td>{{p['miktar']}}</td><td>{{p['birim']}}</td><td>{{p['prim']}}</td><td>{{p['aciklama']}}</td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
  <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
</div>
'''

perf_rapor_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">Performans Raporu</h3>
      <form method="post">
        <div class="mb-3">
          <label>Şantiye:</label>
          <select class="form-select" name="santiye">
            <option value="">Tümü</option>
            {% for s in santiyeler %}
              <option value="{{s['ad']}}">{{s['ad']}}</option>
            {% endfor %}
          </select>
        </div>
        <div class="mb-3">
          <label>Çalışan:</label>
          <select class="form-select" name="calisan">
            <option value="">Tümü</option>
            {% for c in calisanlar %}
              <option value="{{c['ad']}} {{c['soyad']}}">{{c['ad']}} {{c['soyad']}}</option>
            {% endfor %}
          </select>
        </div>
        <div class="mb-3">
          <label>Başlangıç Tarihi:</label>
          <input type="date" class="form-control" name="baslangic">
        </div>
        <div class="mb-3">
          <label>Bitiş Tarihi:</label>
          <input type="date" class="form-control" name="bitis">
        </div>
        <button type="submit" class="btn btn-success w-100">Raporu Oluştur (Excel)</button>
      </form>
      <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
    </div>
  </div>
</div>
'''

sgk_rapor_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">SGK Bildirgesi Raporu</h3>
      <form method="post">
        <div class="mb-3">
          <label>Ay:</label>
          <input type="month" class="form-control" name="ay" required>
        </div>
        <button type="submit" class="btn btn-success w-100">SGK Raporu Oluştur (Excel)</button>
      </form>
      <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
    </div>
  </div>
</div>
'''

@app.route('/sgk_rapor', methods=['GET', 'POST'])
def sgk_rapor():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    calisanlar = load_calisanlar()
    if request.method == 'POST':
        ay = request.form['ay']
        ay_yil = ay.split('-')
        ay_str = ay_yil[0] + '-' + ay_yil[1]
        puantajlar = load_puantajlar()
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Ad Soyad','TC','SGK No','Giriş Tarihi','Ayrılış Tarihi','Ay','Çalışılan Gün'])
        for c in calisanlar:
            puantaj = [p for p in puantajlar if p['calisan']==c['ad']+' '+c['soyad'] and p['tarih'].startswith(ay_str)]
            gun_sayisi = sum(1 for p in puantaj if p['tur']=='Giriş')
            ws.append([
                c['ad']+' '+c['soyad'], c['tc'], c['sgk'], c['giris'], c['ayrilis'], ay, gun_sayisi
            ])
        dosya = f'sgk_rapor_{ay}.xlsx'
        wb.save(dosya)
        from flask import send_file
        return send_file(dosya, as_attachment=True)
    return render_template_string(sgk_rapor_form_page, session=session)

devamsizlik_rapor_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">Devamsızlık/İzin/Fazla Mesai Raporu</h3>
      <form method="post">
        <div class="mb-3">
          <label>Tür:</label>
          <select class="form-select" name="tur" required>
            <option value="Devamsızlık">Devamsızlık</option>
            <option value="İzin">İzin</option>
            <option value="Fazla Mesai">Fazla Mesai</option>
          </select>
        </div>
        <div class="mb-3">
          <label>Ay:</label>
          <input type="month" class="form-control" name="ay" required>
        </div>
        <button type="submit" class="btn btn-success w-100">Raporu Oluştur (Excel)</button>
      </form>
      <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
    </div>
  </div>
</div>
'''

@app.route('/devamsizlik_rapor', methods=['GET', 'POST'])
def devamsizlik_rapor():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    calisanlar = load_calisanlar()
    if request.method == 'POST':
        tur = request.form['tur']
        ay = request.form['ay']
        ay_yil = ay.split('-')
        ay_str = ay_yil[0] + '-' + ay_yil[1]
        puantajlar = load_puantajlar()
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Ad Soyad','Şantiye','Tarih','Saat','Açıklama'])
        for c in calisanlar:
            puantaj = [p for p in puantajlar if p['calisan']==c['ad']+' '+c['soyad'] and p['tur']==tur and p['tarih'].startswith(ay_str)]
            for p in puantaj:
                ws.append([p['calisan'], p['santiye'], p['tarih'], p['saat'], p['aciklama']])
        dosya = f'{tur.lower()}_rapor_{ay}.xlsx'
        wb.save(dosya)
        from flask import send_file
        return send_file(dosya, as_attachment=True)
    return render_template_string(devamsizlik_rapor_form_page, session=session)

# --- Giriş Ekranı ---
@app.route('/', methods=['GET', 'POST'])
def login():
    if 'logged_in' in session:
        return redirect(url_for('menu'))
    error = None
    if request.method == 'POST':
        if request.form['password'] == ADMIN_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('menu'))
        else:
            error = 'Hatalı şifre!'
    return render_template_string(login_page, error=error, session=session)

# --- Ana Menü ---
@app.route('/menu')
def menu():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template_string(menu_page, session=session)

# --- Giriş/Çıkış Kayıtları Listele ---
@app.route('/kayitlar')
def kayitlar():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    data = load_data()
    giris_cikis = [k for k in data if k['tur'] in ['Giriş', 'Çıkış']]
    return render_template_string(kayitlar_page, kayitlar=giris_cikis, session=session)

# --- Giriş/Çıkış Kaydı Ekle ---
@app.route('/kayit_ekle', methods=['GET', 'POST'])
def kayit_ekle():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        data = load_data()
        yeni = {
            'tarih': request.form['tarih'],
            'saat': request.form['saat'],
            'tur': request.form['tur'],
            'aciklama': request.form['aciklama']
        }
        data.append(yeni)
        save_data(data)
        return redirect(url_for('kayitlar'))
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template_string(kayit_ekle_page, today=today, session=session)

# --- İzin/Vardiya/Fazla Mesai Listele ---
@app.route('/ozel')
def ozel():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    data = load_data()
    ozel_kayitlar = [k for k in data if k['tur'] in ['İzin', 'Vardiya', 'Fazla Mesai']]
    return render_template_string(ozel_page, kayitlar=ozel_kayitlar, session=session)

# --- İzin/Vardiya/Fazla Mesai Ekle ---
@app.route('/ozel_ekle', methods=['GET', 'POST'])
def ozel_ekle():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        data = load_data()
        yeni = {
            'tarih': request.form['tarih'],
            'saat': request.form['saat'],
            'tur': request.form['tur'],
            'aciklama': request.form['aciklama']
        }
        data.append(yeni)
        save_data(data)
        return redirect(url_for('ozel'))
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template_string(ozel_ekle_page, today=today, session=session)

# --- Raporlama (Excel İndir) ---
@app.route('/rapor')
def rapor():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    # Excel dosyasını oluştur
    data = load_data()
    wb = Workbook()
    ws = wb.active
    ws.append(['Tarih', 'Saat', 'Tür', 'Açıklama'])
    for k in data:
        ws.append([k['tarih'], k['saat'], k['tur'], k['aciklama']])
    rapor_dosya = 'puantaj_rapor.xlsx'
    wb.save(rapor_dosya)
    return send_file(rapor_dosya, as_attachment=True)

# --- Çalışanlar ---
@app.route('/calisanlar')
def calisanlar():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    calisanlar = load_calisanlar()
    return render_template_string(calisan_list_page, calisanlar=calisanlar, session=session)

@app.route('/calisan_ekle', methods=['GET', 'POST'])
def calisan_ekle():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    santiye_list = load_santiyeler()
    taseron_list = load_taseronlar()
    if request.method == 'POST':
        calisanlar = load_calisanlar()
        yeni = {
            'ad': request.form['ad'],
            'soyad': request.form['soyad'],
            'tc': request.form['tc'],
            'sgk': request.form['sgk'],
            'departman': request.form['departman'],
            'gorev': request.form['gorev'],
            'ustalik': request.form['ustalik'],
            'vardiya': request.form['vardiya'],
            'taseron': request.form['taseron'],
            'santiye': request.form['santiye'],
            'giris': request.form['giris'],
            'ayrilis': request.form['ayrilis'],
            'yevmiye': request.form['yevmiye'],
            'fm_ucret': request.form['fm_ucret'],
            'avans': request.form['avans'],
            'kesinti': request.form['kesinti']
        }
        calisanlar.append(yeni)
        save_calisanlar(calisanlar)
        return redirect(url_for('calisanlar'))
    c = {k: '' for k in ['ad','soyad','tc','sgk','departman','gorev','ustalik','vardiya','taseron','santiye','giris','ayrilis','yevmiye','fm_ucret','avans','kesinti']}
    return render_template_string(calisan_form_page, c=c, baslik='Yeni Çalışan Ekle', santiye_list=santiye_list, taseron_list=taseron_list, session=session)

@app.route('/calisan_duzenle/<int:idx>', methods=['GET', 'POST'])
def calisan_duzenle(idx):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    santiye_list = load_santiyeler()
    taseron_list = load_taseronlar()
    calisanlar = load_calisanlar()
    if idx < 0 or idx >= len(calisanlar):
        return redirect(url_for('calisanlar'))
    if request.method == 'POST':
        calisanlar[idx] = {
            'ad': request.form['ad'],
            'soyad': request.form['soyad'],
            'tc': request.form['tc'],
            'sgk': request.form['sgk'],
            'departman': request.form['departman'],
            'gorev': request.form['gorev'],
            'ustalik': request.form['ustalik'],
            'vardiya': request.form['vardiya'],
            'taseron': request.form['taseron'],
            'santiye': request.form['santiye'],
            'giris': request.form['giris'],
            'ayrilis': request.form['ayrilis'],
            'yevmiye': request.form['yevmiye'],
            'fm_ucret': request.form['fm_ucret'],
            'avans': request.form['avans'],
            'kesinti': request.form['kesinti']
        }
        save_calisanlar(calisanlar)
        return redirect(url_for('calisanlar'))
    c = calisanlar[idx]
    for k in ['yevmiye','fm_ucret','avans','kesinti','taseron']:
        if k not in c:
            c[k] = ''
    return render_template_string(calisan_form_page, c=c, baslik='Çalışan Düzenle', santiye_list=santiye_list, taseron_list=taseron_list, session=session)

@app.route('/calisan_sil/<int:idx>')
def calisan_sil(idx):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    calisanlar = load_calisanlar()
    if 0 <= idx < len(calisanlar):
        calisanlar.pop(idx)
        save_calisanlar(calisanlar)
    return redirect(url_for('calisanlar'))

# --- Şantiyeler ---
@app.route('/santiyeler')
def santiyeler():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    santiyeler = load_santiyeler()
    return render_template_string(santiye_list_page, santiyeler=santiyeler, session=session)

@app.route('/santiye_ekle', methods=['GET', 'POST'])
def santiye_ekle():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        santiyeler = load_santiyeler()
        yeni = {
            'ad': request.form['ad'],
            'lokasyon': request.form['lokasyon'],
            'proje_kodu': request.form['proje_kodu'],
            'yonetici': request.form['yonetici']
        }
        santiyeler.append(yeni)
        save_santiyeler(santiyeler)
        return redirect(url_for('santiyeler'))
    s = {k: '' for k in ['ad','lokasyon','proje_kodu','yonetici']}
    return render_template_string(santiye_form_page, s=s, baslik='Yeni Şantiye Ekle', session=session)

@app.route('/santiye_duzenle/<int:idx>', methods=['GET', 'POST'])
def santiye_duzenle(idx):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    santiyeler = load_santiyeler()
    if idx < 0 or idx >= len(santiyeler):
        return redirect(url_for('santiyeler'))
    if request.method == 'POST':
        santiyeler[idx] = {
            'ad': request.form['ad'],
            'lokasyon': request.form['lokasyon'],
            'proje_kodu': request.form['proje_kodu'],
            'yonetici': request.form['yonetici']
        }
        save_santiyeler(santiyeler)
        return redirect(url_for('santiyeler'))
    s = santiyeler[idx]
    return render_template_string(santiye_form_page, s=s, baslik='Şantiyeyi Düzenle', session=session)

@app.route('/santiye_sil/<int:idx>')
def santiye_sil(idx):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    santiyeler = load_santiyeler()
    if 0 <= idx < len(santiyeler):
        santiyeler.pop(idx)
        save_santiyeler(santiyeler)
    return redirect(url_for('santiyeler'))

# --- Günlük Puantajlar ---
@app.route('/puantajlar')
def puantajlar():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    puantajlar = load_puantajlar()
    return render_template_string(puantaj_list_page, puantajlar=puantajlar, session=session)

@app.route('/puantaj_ekle', methods=['GET', 'POST'])
def puantaj_ekle():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    calisanlar = load_calisanlar()
    santiyeler = load_santiyeler()
    if request.method == 'POST':
        puantajlar = load_puantajlar()
        yeni = {
            'calisan': request.form['calisan'],
            'santiye': request.form['santiye'],
            'tarih': request.form['tarih'],
            'tur': request.form['tur'],
            'saat': request.form['saat'],
            'aciklama': request.form['aciklama']
        }
        puantajlar.append(yeni)
        save_puantajlar(puantajlar)
        return redirect(url_for('puantajlar'))
    kayit = {'calisan':'','santiye':'','tarih':'','tur':'Giriş','saat':'','aciklama':''}
    return render_template_string(puantaj_form_page, calisanlar=calisanlar, santiyeler=santiyeler, kayit=kayit, session=session)

# --- Bordro/Maaş Raporu ---
@app.route('/bordro', methods=['GET', 'POST'])
def bordro():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    calisanlar = load_calisanlar()
    if request.method == 'POST':
        idx = int(request.form['calisan'])
        ay = request.form['ay']  # yyyy-mm
        c = calisanlar[idx]
        puantajlar = load_puantajlar()
        # Seçilen ay ve çalışana göre filtrele
        ay_yil = ay.split('-')
        ay_str = ay_yil[0] + '-' + ay_yil[1]
        puantaj = [p for p in puantajlar if p['calisan']==c['ad']+' '+c['soyad'] and p['tarih'].startswith(ay_str)]
        # Hesaplamalar
        gun_sayisi = sum(1 for p in puantaj if p['tur']=='Giriş')
        izin = sum(1 for p in puantaj if p['tur']=='İzin')
        devamsizlik = sum(1 for p in puantaj if p['tur']=='Devamsızlık')
        fm_saat = sum(float(p['saat']) if p['tur']=='Fazla Mesai' and p['saat'] else 0 for p in puantaj)
        yevmiye = float(c.get('yevmiye',0) or 0)
        fm_ucret = float(c.get('fm_ucret',0) or 0)
        avans = float(c.get('avans',0) or 0)
        kesinti = float(c.get('kesinti',0) or 0)
        toplam_maas = gun_sayisi*yevmiye + fm_saat*fm_ucret - avans - kesinti
        # Excel oluştur
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Ad Soyad','Ay','Gün','İzin','Devamsızlık','Fazla Mesai (saat)','Yevmiye','FM Ücreti','Avans','Kesinti','Net Ödeme'])
        ws.append([
            c['ad']+' '+c['soyad'], ay,
            gun_sayisi, izin, devamsizlik, fm_saat, yevmiye, fm_ucret, avans, kesinti, toplam_maas
        ])
        ws.append([])
        ws.append(['Tarih','Tür','Saat','Açıklama'])
        for p in puantaj:
            ws.append([p['tarih'], p['tur'], p['saat'], p['aciklama']])
        dosya = f'bordro_{c["ad"]}_{c["soyad"]}_{ay}.xlsx'
        wb.save(dosya)
        from flask import send_file
        return send_file(dosya, as_attachment=True)
    return render_template_string(bordro_form_page, calisanlar=calisanlar, session=session)

# --- Çıkış ---
@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

# Taşeron yönetimi rotaları
def get_taseron_list():
    return load_taseronlar()

@app.route('/taseronlar')
def taseronlar():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    taseronlar = load_taseronlar()
    return render_template_string(taseron_list_page, taseronlar=taseronlar, session=session)

@app.route('/taseron_ekle', methods=['GET', 'POST'])
def taseron_ekle():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        taseronlar = load_taseronlar()
        yeni = {
            'ad': request.form['ad'],
            'yetkili': request.form['yetkili'],
            'telefon': request.form['telefon']
        }
        taseronlar.append(yeni)
        save_taseronlar(taseronlar)
        return redirect(url_for('taseronlar'))
    t = {k: '' for k in ['ad','yetkili','telefon']}
    return render_template_string(taseron_form_page, t=t, baslik='Yeni Taşeron Ekle', session=session)

@app.route('/taseron_duzenle/<int:idx>', methods=['GET', 'POST'])
def taseron_duzenle(idx):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    taseronlar = load_taseronlar()
    if idx < 0 or idx >= len(taseronlar):
        return redirect(url_for('taseronlar'))
    if request.method == 'POST':
        taseronlar[idx] = {
            'ad': request.form['ad'],
            'yetkili': request.form['yetkili'],
            'telefon': request.form['telefon']
        }
        save_taseronlar(taseronlar)
        return redirect(url_for('taseronlar'))
    t = taseronlar[idx]
    return render_template_string(taseron_form_page, t=t, baslik='Taşeronu Düzenle', session=session)

@app.route('/taseron_sil/<int:idx>')
def taseron_sil(idx):
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    taseronlar = load_taseronlar()
    if 0 <= idx < len(taseronlar):
        taseronlar.pop(idx)
        save_taseronlar(taseronlar)
    return redirect(url_for('taseronlar'))

# Şantiye bazlı maliyet raporu (Excel)
rapor_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">Şantiye Bazlı Maliyet Raporu</h3>
      <form method="post">
        <div class="mb-3">
          <label>Şantiye:</label>
          <select class="form-select" name="santiye" required>
            <option value="">Şantiye seçiniz</option>
            {% for idx, s in enumerate(santiyeler) %}
              <option value="{{s['ad']}}">{{s['ad']}}</option>
            {% endfor %}
          </select>
        </div>
        <div class="mb-3">
          <label>Ay:</label>
          <input type="month" class="form-control" name="ay" required>
        </div>
        <button type="submit" class="btn btn-success w-100">Raporu Oluştur (Excel)</button>
      </form>
      <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
    </div>
  </div>
</div>
'''

@app.route('/santiye_rapor', methods=['GET', 'POST'])
def santiye_rapor():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    santiyeler = load_santiyeler()
    calisanlar = load_calisanlar()
    if request.method == 'POST':
        santiye = request.form['santiye']
        ay = request.form['ay']
        ay_yil = ay.split('-')
        ay_str = ay_yil[0] + '-' + ay_yil[1]
        puantajlar = load_puantajlar()
        # Şantiyedeki çalışanlar
        calisanlar_s = [c for c in calisanlar if c['santiye']==santiye]
        # Her çalışan için maliyet
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Ad Soyad','Taşeron','Gün','Fazla Mesai (saat)','Yevmiye','FM Ücreti','Avans','Kesinti','Net Ödeme'])
        for c in calisanlar_s:
            puantaj = [p for p in puantajlar if p['calisan']==c['ad']+' '+c['soyad'] and p['santiye']==santiye and p['tarih'].startswith(ay_str)]
            gun_sayisi = sum(1 for p in puantaj if p['tur']=='Giriş')
            fm_saat = sum(float(p['saat']) if p['tur']=='Fazla Mesai' and p['saat'] else 0 for p in puantaj)
            yevmiye = float(c.get('yevmiye',0) or 0)
            fm_ucret = float(c.get('fm_ucret',0) or 0)
            avans = float(c.get('avans',0) or 0)
            kesinti = float(c.get('kesinti',0) or 0)
            toplam_maas = gun_sayisi*yevmiye + fm_saat*fm_ucret - avans - kesinti
            ws.append([
                c['ad']+' '+c['soyad'], c.get('taseron',''), gun_sayisi, fm_saat, yevmiye, fm_ucret, avans, kesinti, toplam_maas
            ])
        dosya = f'santiye_rapor_{santiye}_{ay}.xlsx'
        wb.save(dosya)
        from flask import send_file
        return send_file(dosya, as_attachment=True)
    return render_template_string(rapor_form_page, santiyeler=santiyeler, session=session)

@app.route('/performanslar')
def performanslar():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    performanslar = load_performans()
    return render_template_string(perf_list_page, performanslar=performanslar, session=session)

@app.route('/performans_ekle', methods=['GET', 'POST'])
def performans_ekle():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    calisanlar = load_calisanlar()
    santiyeler = load_santiyeler()
    if request.method == 'POST':
        performanslar = load_performans()
        yeni = {
            'calisan': request.form['calisan'],
            'santiye': request.form['santiye'],
            'tarih': request.form['tarih'],
            'is_turu': request.form['is_turu'],
            'miktar': request.form['miktar'],
            'birim': request.form['birim'],
            'prim': request.form['prim'],
            'aciklama': request.form['aciklama']
        }
        performanslar.append(yeni)
        save_performans(performanslar)
        return redirect(url_for('performanslar'))
    kayit = {'calisan':'','santiye':'','tarih':'','is_turu':'','miktar':'','birim':'','prim':'','aciklama':''}
    return render_template_string(perf_form_page, calisanlar=calisanlar, santiyeler=santiyeler, kayit=kayit, session=session)

@app.route('/performans_rapor', methods=['GET', 'POST'])
def performans_rapor():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    calisanlar = load_calisanlar()
    santiyeler = load_santiyeler()
    if request.method == 'POST':
        calisan = request.form['calisan']
        santiye = request.form['santiye']
        baslangic = request.form['baslangic']
        bitis = request.form['bitis']
        performanslar = load_performans()
        # Filtrele
        filtered = []
        for p in performanslar:
            if calisan and p['calisan'] != calisan:
                continue
            if santiye and p['santiye'] != santiye:
                continue
            if baslangic and p['tarih'] < baslangic:
                continue
            if bitis and p['tarih'] > bitis:
                continue
            filtered.append(p)
        # Excel oluştur
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Çalışan','Şantiye','Tarih','İş Türü','Miktar','Birim','Prim/Ödül','Açıklama'])
        for p in filtered:
            ws.append([p['calisan'],p['santiye'],p['tarih'],p['is_turu'],p['miktar'],p['birim'],p['prim'],p['aciklama']])
        dosya = 'performans_rapor.xlsx'
        wb.save(dosya)
        from flask import send_file
        return send_file(dosya, as_attachment=True)
    return render_template_string(perf_rapor_form_page, calisanlar=calisanlar, santiyeler=santiyeler, session=session)

toplu_bordro_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">Toplu Aylık Bordro Raporu</h3>
      <form method="post">
        <div class="mb-3">
          <label>Ay:</label>
          <input type="month" class="form-control" name="ay" required>
        </div>
        <button type="submit" class="btn btn-success w-100">Toplu Bordro Oluştur (Excel)</button>
      </form>
      <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
    </div>
  </div>
</div>
'''

@app.route('/toplu_bordro', methods=['GET', 'POST'])
def toplu_bordro():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    calisanlar = load_calisanlar()
    if request.method == 'POST':
        ay = request.form['ay']
        ay_yil = ay.split('-')
        ay_str = ay_yil[0] + '-' + ay_yil[1]
        puantajlar = load_puantajlar()
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Ad Soyad','Şantiye','Taşeron','Gün','Fazla Mesai (saat)','Yevmiye','FM Ücreti','Avans','Kesinti','Net Ödeme'])
        for c in calisanlar:
            puantaj = [p for p in puantajlar if p['calisan']==c['ad']+' '+c['soyad'] and p['tarih'].startswith(ay_str)]
            gun_sayisi = sum(1 for p in puantaj if p['tur']=='Giriş')
            fm_saat = sum(float(p['saat']) if p['tur']=='Fazla Mesai' and p['saat'] else 0 for p in puantaj)
            yevmiye = float(c.get('yevmiye',0) or 0)
            fm_ucret = float(c.get('fm_ucret',0) or 0)
            avans = float(c.get('avans',0) or 0)
            kesinti = float(c.get('kesinti',0) or 0)
            toplam_maas = gun_sayisi*yevmiye + fm_saat*fm_ucret - avans - kesinti
            ws.append([
                c['ad']+' '+c['soyad'], c.get('santiye',''), c.get('taseron',''), gun_sayisi, fm_saat, yevmiye, fm_ucret, avans, kesinti, toplam_maas
            ])
        dosya = f'toplu_bordro_{ay}.xlsx'
        wb.save(dosya)
        from flask import send_file
        return send_file(dosya, as_attachment=True)
    return render_template_string(toplu_bordro_form_page, session=session)

taseron_rapor_form_page = BOOTSTRAP + HEADER + '''
<div class="container" style="max-width:500px;">
  <div class="card shadow">
    <div class="card-body">
      <h3 class="mb-4">Taşeron Bazlı Maliyet Raporu</h3>
      <form method="post">
        <div class="mb-3">
          <label>Taşeron:</label>
          <select class="form-select" name="taseron" required>
            <option value="">Taşeron seçiniz</option>
            {% for t in taseronlar %}
              <option value="{{t['ad']}}">{{t['ad']}}</option>
            {% endfor %}
          </select>
        </div>
        <div class="mb-3">
          <label>Ay:</label>
          <input type="month" class="form-control" name="ay" required>
        </div>
        <button type="submit" class="btn btn-success w-100">Raporu Oluştur (Excel)</button>
      </form>
      <a href="/menu" class="btn btn-secondary mt-3">Ana Menü</a>
    </div>
  </div>
</div>
'''

@app.route('/taseron_rapor', methods=['GET', 'POST'])
def taseron_rapor():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    taseronlar = load_taseronlar()
    calisanlar = load_calisanlar()
    if request.method == 'POST':
        taseron = request.form['taseron']
        ay = request.form['ay']
        ay_yil = ay.split('-')
        ay_str = ay_yil[0] + '-' + ay_yil[1]
        puantajlar = load_puantajlar()
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['Ad Soyad','Şantiye','Gün','Fazla Mesai (saat)','Yevmiye','FM Ücreti','Avans','Kesinti','Net Ödeme'])
        for c in calisanlar:
            if c.get('taseron','') != taseron:
                continue
            puantaj = [p for p in puantajlar if p['calisan']==c['ad']+' '+c['soyad'] and p['tarih'].startswith(ay_str)]
            gun_sayisi = sum(1 for p in puantaj if p['tur']=='Giriş')
            fm_saat = sum(float(p['saat']) if p['tur']=='Fazla Mesai' and p['saat'] else 0 for p in puantaj)
            yevmiye = float(c.get('yevmiye',0) or 0)
            fm_ucret = float(c.get('fm_ucret',0) or 0)
            avans = float(c.get('avans',0) or 0)
            kesinti = float(c.get('kesinti',0) or 0)
            toplam_maas = gun_sayisi*yevmiye + fm_saat*fm_ucret - avans - kesinti
            ws.append([
                c['ad']+' '+c['soyad'], c.get('santiye',''), gun_sayisi, fm_saat, yevmiye, fm_ucret, avans, kesinti, toplam_maas
            ])
        dosya = f'taseron_rapor_{taseron}_{ay}.xlsx'
        wb.save(dosya)
        from flask import send_file
        return send_file(dosya, as_attachment=True)
    return render_template_string(taseron_rapor_form_page, taseronlar=taseronlar, session=session)

@app.route('/', methods=['GET', 'POST'])
def login():
    error = None
    if 'logged_in' in session:
        return redirect(url_for('menu'))
    if request.method == 'POST':
        if request.form['password'] == ADMIN_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('menu'))
        else:
            error = 'Hatalı şifre!'
    return render_template_string(login_page, error=error, session=session)

@app.route('/menu')
def menu():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return render_template_string(menu_page, session=session)

if __name__ == '__main__':
    app.run(debug=True)
